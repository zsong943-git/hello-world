"""
canteen_tool.py
===============
华东食堂出入库及盘点表一体化自动填写工具

包含三大功能模块：
  1. 入库单（ruku）  —— 将「食材每日出入库.xlsx」入库数据填入「华东XXXX年X月入库单.xlsx」
  2. 出库单（chuku） —— 将「食材每日出入库.xlsx」出库数据填入「华东XXXX年X月出库单.xlsx」
  3. 存货盘点表（pandian） —— 将「食材每日出入库.xlsx」汇总数据填入「华东2026年X月存货盘点表.xlsx」

用法：
    python canteen_tool.py                         # 交互式菜单
    python canteen_tool.py ruku    --month 1       # 直接处理1月入库单
    python canteen_tool.py chuku   --month 1       # 直接处理1月出库单
    python canteen_tool.py pandian --month 1       # 直接处理1月存货盘点表
    python canteen_tool.py all     --month 1       # 同时处理三个模块
"""

import os
import argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles import numbers as xl_numbers


# ══════════════════════════════════════════════════════════════
# 共用全局配置
# ══════════════════════════════════════════════════════════════

SOURCE_FILE = "2026年01-食材每日出入库.xlsx"   # 数据源文件（三个模块共用）

# 各类别固定排列顺序（入库单、出库单、盘点表共用）
CATEGORY_ORDER = ["蔬菜", "水果", "蛋", "豆制品", "米面（粮）", "油", "调料", "半成品", "肉", "其他"]

# 入库单月份配置
RUKU_MONTH_CONFIG = {
    m: {"target_file": f"华东2026年{m}月入库单.xlsx", "year": 2026, "month": m}
    for m in range(1, 13)
}

# 出库单月份配置
CHUKU_MONTH_CONFIG = {
    m: {"target_file": f"华东2026年{m}月出库单.xlsx", "year": 2026, "month": m}
    for m in range(1, 13)
}

# 盘点表目标文件命名函数
def pandian_target_file(month):
    return f"华东2026年{month}月存货盘点表.xlsx"

# 盘点表需要新建的四张明细表：(sheet名, 货品系列名, 盘点类别显示名)
DETAIL_SHEETS = [
    ("米面粮明细表", "米面（粮）", "米面粮"),
    ("油明细表",     "油",        "油"),
    ("调料明细表",   "调料",      "调料"),
    ("半成品明细表", "半成品",    "半成品"),
]

# 数字格式：保留两位小数
FMT_NUMBER = "0.00"


# ══════════════════════════════════════════════════════════════
# 共用样式辅助 —— Side 常量
# ══════════════════════════════════════════════════════════════

_THIN   = Side(style="thin")
_MEDIUM = Side(style="medium")
_NONE   = Side(style=None)


def _border(left=None, right=None, top=None, bottom=None):
    """通用边框构造器。"""
    return Border(
        left   = left   if left   is not None else _THIN,
        right  = right  if right  is not None else _THIN,
        top    = top    if top    is not None else _THIN,
        bottom = bottom if bottom is not None else _THIN,
    )


def mk_font(bold=False, size=10, name="宋体"):
    """创建字体对象。"""
    return Font(bold=bold, size=size, name=name)


def mk_align(h="center", v="center"):
    """创建对齐对象，默认水平垂直居中。"""
    return Alignment(horizontal=h, vertical=v)


def thin_border():
    """四边 thin 的标准边框。"""
    return Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ══════════════════════════════════════════════════════════════
# ██████████████████  模块一：入库单  ██████████████████
# ══════════════════════════════════════════════════════════════

# 入库单各列列宽（A~I，共9列）
RUKU_COL_WIDTHS = {
    "A": 9.4,   # 类别
    "B": 6.5,   # 序号
    "C": 16.6,  # 名称
    "D": 7.6,   # 单位
    "E": 9.0,   # 数量
    "F": 10.8,  # 单价（元）
    "G": 13.6,  # 金额（元）
    "H": 10.0,  # 品质标准
    "I": 9.0,   # 供应商
}

# ── 入库单边框函数 ──

def ruku_border_header_row(col):
    """入库单表头行边框（第3行）：顶部 medium，A列左边 medium，H/I列右边 medium。"""
    left  = _MEDIUM if col == 1 else _THIN
    right = _MEDIUM if col in (8, 9) else _THIN
    return Border(left=left, right=right, top=_MEDIUM, bottom=_THIN)


def ruku_border_data_row(col):
    """入库单普通数据行边框：A列左边 medium，H/I列右边 medium，其余 thin。"""
    left  = _MEDIUM if col == 1 else _THIN
    right = _MEDIUM if col in (8, 9) else _THIN
    return Border(left=left, right=right, top=_THIN, bottom=_THIN)


def ruku_border_subtotal_row(col):
    """入库单金额合计行边框：底部全列 medium，合并区内部列无竖线。"""
    bottom = _MEDIUM
    top    = _THIN
    if col == 1:
        return Border(left=_MEDIUM, right=_THIN,   top=top, bottom=bottom)
    elif col == 2:
        return Border(left=_THIN,   right=_THIN,   top=top, bottom=bottom)
    elif col in (3, 4, 5):
        return Border(left=_NONE,   right=_NONE,   top=top, bottom=bottom)
    elif col == 6:
        return Border(left=_NONE,   right=_THIN,   top=top, bottom=bottom)
    elif col == 7:
        return Border(left=_THIN,   right=_THIN,   top=top, bottom=bottom)
    elif col == 8:
        return Border(left=_THIN,   right=_MEDIUM, top=top, bottom=bottom)
    else:  # col == 9
        return Border(left=_MEDIUM, right=_MEDIUM, top=top, bottom=bottom)


def ruku_border_total_row(col):
    """入库单总计金额行边框：底部全列 medium，顶部 None。"""
    bottom = _MEDIUM
    top    = _NONE
    if col == 1:
        return Border(left=_MEDIUM, right=_THIN,   top=top, bottom=bottom)
    elif col in (2, 3, 4, 5):
        return Border(left=_NONE,   right=_NONE,   top=top, bottom=bottom)
    elif col == 6:
        return Border(left=_NONE,   right=_THIN,   top=top, bottom=bottom)
    elif col == 7:
        return Border(left=_THIN,   right=_THIN,   top=top, bottom=bottom)
    elif col == 8:
        return Border(left=_THIN,   right=_MEDIUM, top=top, bottom=bottom)
    else:  # col == 9
        return Border(left=_NONE,   right=_MEDIUM, top=top, bottom=bottom)


def ruku_set_cell(ws, row, col, value, bold=False, size=10, h_align="center",
                  border_fn=None, font_name="宋体"):
    """向入库单单元格写入值并设置样式。"""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = mk_font(bold=bold, size=size, name=font_name)
    c.alignment = mk_align(h=h_align)
    if border_fn is not None:
        c.border = border_fn(col)
    return c


def ruku_write_day_sheet(ws, date_ts, day_df, sign_rows):
    """
    向工作表 ws 写入某一天完整的入库单内容（9列：含供应商）。
    """
    year  = date_ts.year
    month = date_ts.month
    day   = date_ts.day

    for col_letter, w in RUKU_COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = w

    # 第1行：标题
    ws.merge_cells("A1:I1")
    c1 = ws["A1"]
    c1.value     = "华东师大泰达附属学校食堂入库单"
    c1.font      = mk_font(bold=True, size=18)
    c1.alignment = mk_align(h="center")
    ws.row_dimensions[1].height = 26

    # 第2行：日期
    ws.merge_cells("A2:E2")
    c2 = ws["A2"]
    c2.value     = f"日期：{year}/{month}/{day}"
    c2.font      = mk_font(bold=True, size=12)
    c2.alignment = mk_align(h="left")
    ws.row_dimensions[2].height = 20

    # 第3行：表头
    headers = ["类别", "序号", "名称", "单位", "数量", "单价（元）", "金额（元）", "品质标准", "供应商"]
    for col_i, h_text in enumerate(headers, 1):
        ruku_set_cell(ws, 3, col_i, h_text, bold=True, size=12,
                      border_fn=ruku_border_header_row)
    ws.row_dimensions[3].height = 15.5

    # 数据区
    current_row   = 4
    subtotal_refs = []

    for cat in CATEGORY_ORDER:
        cat_df = day_df[day_df["货品系列"] == cat].reset_index(drop=True)
        cat_start_row = current_row

        if len(cat_df) == 0:
            r = current_row
            ruku_set_cell(ws, r, 1, cat, size=10, border_fn=ruku_border_data_row)
            ruku_set_cell(ws, r, 2, 1,   size=10, border_fn=ruku_border_data_row)
            for col_i in range(3, 10):
                ruku_set_cell(ws, r, col_i, None, size=10, border_fn=ruku_border_data_row)
            ws.cell(row=r, column=7).value = f"=ROUND(E{r}*F{r},2)"
            ws.row_dimensions[r].height = 15.5
            current_row += 1
        else:
            for seq, (_, row_data) in enumerate(cat_df.iterrows(), 1):
                r = current_row
                ruku_set_cell(ws, r, 1, cat,                  size=10, border_fn=ruku_border_data_row)
                ruku_set_cell(ws, r, 2, seq,                  size=10, border_fn=ruku_border_data_row)
                ruku_set_cell(ws, r, 3, row_data["货品名称"],  size=10, border_fn=ruku_border_data_row)
                ruku_set_cell(ws, r, 4, row_data["单位"],      size=10, border_fn=ruku_border_data_row)
                ruku_set_cell(ws, r, 5, row_data["数量"],      size=10, border_fn=ruku_border_data_row)
                ruku_set_cell(ws, r, 6, row_data["单价"],      size=10, border_fn=ruku_border_data_row)
                ws.cell(row=r, column=7).value  = f"=ROUND(E{r}*F{r},2)"
                ws.cell(row=r, column=7).font   = mk_font(size=10)
                ws.cell(row=r, column=7).border = ruku_border_data_row(7)
                ruku_set_cell(ws, r, 8, "良好",                size=10, border_fn=ruku_border_data_row)
                ruku_set_cell(ws, r, 9, row_data["供应商"],    size=10, border_fn=ruku_border_data_row)
                ws.row_dimensions[r].height = 15.5
                current_row += 1

        # 金额合计行
        r = current_row
        data_range = f"G{cat_start_row}:G{current_row - 1}"
        ws.cell(row=r, column=1).border = ruku_border_subtotal_row(1)
        ws.merge_cells(f"B{r}:F{r}")
        c_label = ws.cell(row=r, column=2, value="金额合计：")
        c_label.font      = mk_font(bold=True, size=12)
        c_label.alignment = mk_align(h="center")
        c_label.border    = ruku_border_subtotal_row(2)
        for col_i in (3, 4, 5, 6):
            ws.cell(row=r, column=col_i).border = ruku_border_subtotal_row(col_i)
        cg = ws.cell(row=r, column=7, value=f"=SUM({data_range})")
        cg.font   = mk_font(bold=True, size=12)
        cg.border = ruku_border_subtotal_row(7)
        ws.cell(row=r, column=8).border = ruku_border_subtotal_row(8)
        ws.cell(row=r, column=9).border = ruku_border_subtotal_row(9)
        ws.row_dimensions[r].height = 15.5
        subtotal_refs.append(f"G{r}")
        current_row += 1

    # 总计金额行
    total_row = current_row
    ws.merge_cells(f"A{total_row}:F{total_row}")
    ct = ws.cell(row=total_row, column=1, value="总计金额")
    ct.font      = mk_font(bold=True, size=12)
    ct.alignment = mk_align(h="center")
    ct.border    = ruku_border_total_row(1)
    for col_i in (2, 3, 4, 5, 6):
        ws.cell(row=total_row, column=col_i).border = ruku_border_total_row(col_i)
    total_formula = "=" + "+".join(subtotal_refs)
    cgt = ws.cell(row=total_row, column=7, value=total_formula)
    cgt.font   = mk_font(bold=True, size=12)
    cgt.border = ruku_border_total_row(7)
    ws.cell(row=total_row, column=8).border = ruku_border_total_row(8)
    ws.cell(row=total_row, column=9).border = ruku_border_total_row(9)
    ws.row_dimensions[total_row].height = 15.5
    current_row += 1

    # 空行
    current_row += 1

    # 签名行
    for (left_val, right_val) in sign_rows:
        r = current_row
        if left_val:
            ws.merge_cells(f"A{r}:B{r}")
            cs = ws.cell(row=r, column=1, value=left_val)
            cs.font      = mk_font(size=11)
            cs.alignment = mk_align(h="center")
        if right_val:
            ws.merge_cells(f"E{r}:F{r}")
            cs2 = ws.cell(row=r, column=5, value=right_val)
            cs2.font      = mk_font(size=11)
            cs2.alignment = mk_align(h="center")
        current_row += 1


def process_ruku(month_num):
    """处理指定月份的入库单。"""
    if month_num not in RUKU_MONTH_CONFIG:
        print(f"❌ 不支持的月份：{month_num}，请输入 1~12。")
        return

    cfg         = RUKU_MONTH_CONFIG[month_num]
    target_file = cfg["target_file"]
    year        = cfg["year"]
    month       = cfg["month"]

    if not os.path.exists(SOURCE_FILE):
        print(f"❌ 数据源文件不存在：{SOURCE_FILE}")
        return
    if not os.path.exists(target_file):
        print(f"❌ 目标文件不存在：{target_file}（请先准备好该月份的模板文件）")
        return

    print(f"📂 [入库单] 读取数据源：{SOURCE_FILE}  →  Sheet: 入库数据")
    df_raw = pd.read_excel(SOURCE_FILE, sheet_name="入库数据", header=0)
    df_raw.columns = ["日期", "代码", "货品系列", "货品名称", "规格型号",
                      "单位", "数量", "单价", "金额", "供应商"]
    df_raw["日期"] = pd.to_datetime(df_raw["日期"], errors="coerce")
    df_raw = df_raw.dropna(subset=["日期"])
    mask     = (df_raw["日期"].dt.year == year) & (df_raw["日期"].dt.month == month)
    df_month = df_raw[mask].copy()

    if df_month.empty:
        print(f"⚠️  未找到 {year} 年 {month} 月的入库记录，跳过。")
        return

    unique_dates = sorted(df_month["日期"].unique())
    print(f"✅ 找到 {year}/{month} 月入库记录，共 {len(unique_dates)} 个日期，{len(df_month)} 行数据。")

    print(f"📂 打开目标文件：{target_file}")
    wb = load_workbook(target_file)
    ws_summary = wb["汇总表"]
    sign_rows  = []
    for row in ws_summary.iter_rows():
        for cell in row:
            val = cell.value
            if val and ("经手人" in str(val) or "食堂负责人" in str(val)):
                e_val = ws_summary.cell(row=cell.row, column=5).value
                sign_rows.append((val, e_val or ""))
                break
    while len(sign_rows) < 3:
        sign_rows.append(("", ""))

    for sname in wb.sheetnames[:]:
        if sname != "汇总表":
            del wb[sname]
            print(f"  🗑  删除旧 sheet：{sname}")

    for date_ts in unique_dates:
        m = date_ts.month
        d = date_ts.day
        sheet_name = f"{m}.{d}"
        ws_new = wb.create_sheet(title=sheet_name)
        print(f"  📋 新建 sheet：{sheet_name}  ({year}/{m}/{d})")
        day_df = df_month[df_month["日期"] == date_ts].copy()
        ruku_write_day_sheet(ws_new, date_ts, day_df, sign_rows)

    wb.save(target_file)
    sheet_list = [f"{pd.Timestamp(d).month}.{pd.Timestamp(d).day}" for d in unique_dates]
    print(f"\n✅ [入库单] 完成！已保存：{target_file}")
    print(f"   共生成 {len(unique_dates)} 个日期 sheet：{sheet_list}")


# ══════════════════════════════════════════════════════════════
# ██████████████████  模块二：出库单  ██████████████████
# ══════════════════════════════════════════════════════════════

# 出库单各列列宽（A~H，共8列，无供应商）
CHUKU_COL_WIDTHS = {
    "A": 9.4,   # 类别
    "B": 6.5,   # 序号
    "C": 16.6,  # 名称
    "D": 7.6,   # 单位
    "E": 9.0,   # 数量
    "F": 10.8,  # 单价（元）
    "G": 13.6,  # 金额（元）
    "H": 10.0,  # 品质标准（最后一列）
}

# ── 出库单边框函数 ──

def chuku_border_header_row(col):
    """出库单表头行边框：顶部 medium，A列左边 medium，H列右边 medium。"""
    left  = _MEDIUM if col == 1 else _THIN
    right = _MEDIUM if col == 8 else _THIN
    return Border(left=left, right=right, top=_MEDIUM, bottom=_THIN)


def chuku_border_data_row(col):
    """出库单普通数据行边框：A列左边 medium，H列右边 medium，其余 thin。"""
    left  = _MEDIUM if col == 1 else _THIN
    right = _MEDIUM if col == 8 else _THIN
    return Border(left=left, right=right, top=_THIN, bottom=_THIN)


def chuku_border_subtotal_row(col):
    """出库单金额合计行边框：底部全列 medium，合并区内部列无竖线。"""
    bottom = _MEDIUM
    top    = _THIN
    if col == 1:
        return Border(left=_MEDIUM, right=_THIN,   top=top, bottom=bottom)
    elif col == 2:
        return Border(left=_THIN,   right=_THIN,   top=top, bottom=bottom)
    elif col in (3, 4, 5):
        return Border(left=_NONE,   right=_NONE,   top=top, bottom=bottom)
    elif col == 6:
        return Border(left=_NONE,   right=_THIN,   top=top, bottom=bottom)
    elif col == 7:
        return Border(left=_THIN,   right=_THIN,   top=top, bottom=bottom)
    else:  # col == 8
        return Border(left=_THIN,   right=_MEDIUM, top=top, bottom=bottom)


def chuku_border_total_row(col):
    """出库单总计金额行边框：底部全列 medium，顶部 None。"""
    bottom = _MEDIUM
    top    = _NONE
    if col == 1:
        return Border(left=_MEDIUM, right=_THIN,   top=top, bottom=bottom)
    elif col in (2, 3, 4, 5):
        return Border(left=_NONE,   right=_NONE,   top=top, bottom=bottom)
    elif col == 6:
        return Border(left=_NONE,   right=_THIN,   top=top, bottom=bottom)
    elif col == 7:
        return Border(left=_THIN,   right=_THIN,   top=top, bottom=bottom)
    else:  # col == 8
        return Border(left=_THIN,   right=_MEDIUM, top=top, bottom=bottom)


def chuku_set_cell(ws, row, col, value, bold=False, size=10, h_align="center",
                   border_fn=None, font_name="宋体"):
    """向出库单单元格写入值并设置样式。"""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = mk_font(bold=bold, size=size, name=font_name)
    c.alignment = mk_align(h=h_align)
    if border_fn is not None:
        c.border = border_fn(col)
    return c


def chuku_write_day_sheet(ws, date_ts, day_df, sign_rows):
    """
    向工作表 ws 写入某一天完整的出库单内容（8列：无供应商）。
    """
    year  = date_ts.year
    month = date_ts.month
    day   = date_ts.day

    for col_letter, w in CHUKU_COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = w

    # 第1行：标题
    ws.merge_cells("A1:H1")
    c1 = ws["A1"]
    c1.value     = "华东师大泰达附属学校食堂出库单"
    c1.font      = mk_font(bold=True, size=18)
    c1.alignment = mk_align(h="center")
    ws.row_dimensions[1].height = 26

    # 第2行：日期
    ws.merge_cells("A2:E2")
    c2 = ws["A2"]
    c2.value     = f"日期：{year}/{month}/{day}"
    c2.font      = mk_font(bold=True, size=12)
    c2.alignment = mk_align(h="left")
    ws.row_dimensions[2].height = 20

    # 第3行：表头（8列，无供应商）
    headers = ["类别", "序号", "名称", "单位", "数量", "单价（元）", "金额（元）", "品质标准"]
    for col_i, h_text in enumerate(headers, 1):
        chuku_set_cell(ws, 3, col_i, h_text, bold=True, size=12,
                       border_fn=chuku_border_header_row)
    ws.row_dimensions[3].height = 15.5

    # 数据区
    current_row   = 4
    subtotal_refs = []

    for cat in CATEGORY_ORDER:
        cat_df = day_df[day_df["货品系列"] == cat].reset_index(drop=True)
        cat_start_row = current_row

        if len(cat_df) == 0:
            r = current_row
            chuku_set_cell(ws, r, 1, cat, size=10, border_fn=chuku_border_data_row)
            chuku_set_cell(ws, r, 2, 1,   size=10, border_fn=chuku_border_data_row)
            for col_i in range(3, 9):
                chuku_set_cell(ws, r, col_i, None, size=10, border_fn=chuku_border_data_row)
            ws.cell(row=r, column=7).value = f"=ROUND(E{r}*F{r},2)"
            ws.row_dimensions[r].height = 15.5
            current_row += 1
        else:
            for seq, (_, row_data) in enumerate(cat_df.iterrows(), 1):
                r = current_row
                chuku_set_cell(ws, r, 1, cat,                  size=10, border_fn=chuku_border_data_row)
                chuku_set_cell(ws, r, 2, seq,                  size=10, border_fn=chuku_border_data_row)
                chuku_set_cell(ws, r, 3, row_data["货品名称"],  size=10, border_fn=chuku_border_data_row)
                chuku_set_cell(ws, r, 4, row_data["单位"],      size=10, border_fn=chuku_border_data_row)
                chuku_set_cell(ws, r, 5, row_data["数量"],      size=10, border_fn=chuku_border_data_row)
                chuku_set_cell(ws, r, 6, row_data["单价"],      size=10, border_fn=chuku_border_data_row)
                ws.cell(row=r, column=7).value  = f"=ROUND(E{r}*F{r},2)"
                ws.cell(row=r, column=7).font   = mk_font(size=10)
                ws.cell(row=r, column=7).border = chuku_border_data_row(7)
                chuku_set_cell(ws, r, 8, "良好", size=10, border_fn=chuku_border_data_row)
                ws.row_dimensions[r].height = 15.5
                current_row += 1

        # 金额合计行
        r = current_row
        data_range = f"G{cat_start_row}:G{current_row - 1}"
        ws.cell(row=r, column=1).border = chuku_border_subtotal_row(1)
        ws.merge_cells(f"B{r}:F{r}")
        c_label = ws.cell(row=r, column=2, value="金额合计：")
        c_label.font      = mk_font(bold=True, size=12)
        c_label.alignment = mk_align(h="center")
        c_label.border    = chuku_border_subtotal_row(2)
        for col_i in (3, 4, 5, 6):
            ws.cell(row=r, column=col_i).border = chuku_border_subtotal_row(col_i)
        cg = ws.cell(row=r, column=7, value=f"=SUM({data_range})")
        cg.font   = mk_font(bold=True, size=12)
        cg.border = chuku_border_subtotal_row(7)
        ws.cell(row=r, column=8).border = chuku_border_subtotal_row(8)
        ws.row_dimensions[r].height = 15.5
        subtotal_refs.append(f"G{r}")
        current_row += 1

    # 总计金额行
    total_row = current_row
    ws.merge_cells(f"A{total_row}:F{total_row}")
    ct = ws.cell(row=total_row, column=1, value="总计金额")
    ct.font      = mk_font(bold=True, size=12)
    ct.alignment = mk_align(h="center")
    ct.border    = chuku_border_total_row(1)
    for col_i in (2, 3, 4, 5, 6):
        ws.cell(row=total_row, column=col_i).border = chuku_border_total_row(col_i)
    total_formula = "=" + "+".join(subtotal_refs)
    cgt = ws.cell(row=total_row, column=7, value=total_formula)
    cgt.font   = mk_font(bold=True, size=12)
    cgt.border = chuku_border_total_row(7)
    ws.cell(row=total_row, column=8).border = chuku_border_total_row(8)
    ws.row_dimensions[total_row].height = 15.5
    current_row += 1

    # 空行
    current_row += 1

    # 签名行
    for (left_val, right_val) in sign_rows:
        r = current_row
        if left_val:
            ws.merge_cells(f"A{r}:B{r}")
            cs = ws.cell(row=r, column=1, value=left_val)
            cs.font      = mk_font(size=11)
            cs.alignment = mk_align(h="center")
        if right_val:
            ws.merge_cells(f"E{r}:F{r}")
            cs2 = ws.cell(row=r, column=5, value=right_val)
            cs2.font      = mk_font(size=11)
            cs2.alignment = mk_align(h="center")
        current_row += 1


def process_chuku(month_num):
    """处理指定月份的出库单。"""
    if month_num not in CHUKU_MONTH_CONFIG:
        print(f"❌ 不支持的月份：{month_num}，请输入 1~12。")
        return

    cfg         = CHUKU_MONTH_CONFIG[month_num]
    target_file = cfg["target_file"]
    year        = cfg["year"]
    month       = cfg["month"]

    if not os.path.exists(SOURCE_FILE):
        print(f"❌ 数据源文件不存在：{SOURCE_FILE}")
        return
    if not os.path.exists(target_file):
        print(f"❌ 目标文件不存在：{target_file}（请先准备好该月份的模板文件）")
        return

    print(f"📂 [出库单] 读取数据源：{SOURCE_FILE}  →  Sheet: 出库数据")
    df_raw = pd.read_excel(SOURCE_FILE, sheet_name="出库数据", header=0)
    df_raw = df_raw.iloc[:, :9]
    df_raw.columns = ["日期", "代码", "货品系列", "货品名称", "规格型号",
                      "单位", "数量", "单价", "金额"]
    df_raw["日期"] = pd.to_datetime(df_raw["日期"], errors="coerce")
    df_raw = df_raw.dropna(subset=["日期"])
    mask     = (df_raw["日期"].dt.year == year) & (df_raw["日期"].dt.month == month)
    df_month = df_raw[mask].copy()

    if df_month.empty:
        print(f"⚠️  未找到 {year} 年 {month} 月的出库记录，跳过。")
        return

    unique_dates = sorted(df_month["日期"].unique())
    print(f"✅ 找到 {year}/{month} 月出库记录，共 {len(unique_dates)} 个日期，{len(df_month)} 行数据。")

    print(f"📂 打开目标文件：{target_file}")
    wb = load_workbook(target_file)
    ws_summary = wb["汇总表"]
    sign_rows  = []
    for row in ws_summary.iter_rows():
        for cell in row:
            val = cell.value
            if val and ("经手人" in str(val) or "食堂负责人" in str(val)):
                e_val = ws_summary.cell(row=cell.row, column=5).value
                sign_rows.append((val, e_val or ""))
                break
    while len(sign_rows) < 3:
        sign_rows.append(("", ""))

    for sname in wb.sheetnames[:]:
        if sname != "汇总表":
            del wb[sname]
            print(f"  🗑  删除旧 sheet：{sname}")

    for date_ts in unique_dates:
        m = date_ts.month
        d = date_ts.day
        sheet_name = f"{m}.{d}"
        ws_new = wb.create_sheet(title=sheet_name)
        print(f"  📋 新建 sheet：{sheet_name}  ({year}/{m}/{d})")
        day_df = df_month[df_month["日期"] == date_ts].copy()
        chuku_write_day_sheet(ws_new, date_ts, day_df, sign_rows)

    wb.save(target_file)
    sheet_list = [f"{pd.Timestamp(d).month}.{pd.Timestamp(d).day}" for d in unique_dates]
    print(f"\n✅ [出库单] 完成！已保存：{target_file}")
    print(f"   共生成 {len(unique_dates)} 个日期 sheet：{sheet_list}")


# ══════════════════════════════════════════════════════════════
# ████████████████  模块三：存货盘点表  ████████████████
# ══════════════════════════════════════════════════════════════

def pandian_sc(ws, row, col, value, bold=False, size=12, h="center",
               border=True, fmt=None):
    """向盘点表单元格写入值，统一设置字体、对齐、边框和数字格式。"""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = mk_font(bold=bold, size=size)
    c.alignment = mk_align(h=h)
    if border:
        c.border = thin_border()
    if fmt is not None:
        c.number_format = fmt
    return c


def load_summary_data():
    """
    读取源文件"汇总" sheet 的实际计算值（data_only=True）。

    源表列位置（0-indexed）：
        B(1)=货品系列, C(2)=货品名称
        F(5)=期初数量,  H(7)=期初金额
        I(8)=入库数量,  K(10)=入库金额
        L(11)=出库数量, M(12)=出库金额
    """
    wb = load_workbook(SOURCE_FILE, data_only=True)
    ws = wb["汇总"]

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[1]:
            continue
        rows.append({
            "货品系列": row[1],
            "货品名称": row[2],
            "期初数量": row[5]  or 0,
            "期初金额": row[7]  or 0,
            "入库数量": row[8]  or 0,
            "入库金额": row[10] or 0,
            "出库数量": row[11] or 0,
            "出库金额": row[12] or 0,
        })

    df = pd.DataFrame(rows)
    num_cols = ["期初数量", "期初金额", "入库数量", "入库金额", "出库数量", "出库金额"]
    for col in num_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def pandian_update_summary_sheet(ws, df, month):
    """
    更新已有汇总表的数据区，保留所有原有样式和公式。
    """
    ws["A2"].value = f"盘点期间：2026年{month}月"

    grouped = df.groupby("货品系列").agg(
        期初数量=("期初数量", "sum"),
        期初金额=("期初金额", "sum"),
        入库数量=("入库数量", "sum"),
        入库金额=("入库金额", "sum"),
        出库数量=("出库数量", "sum"),
        出库金额=("出库金额", "sum"),
    )

    for i, cat in enumerate(CATEGORY_ORDER):
        row = 5 + i
        if cat in grouped.index:
            g = grouped.loc[cat]
            ws.cell(row, 3).value = round(float(g["期初数量"]), 2)
            ws.cell(row, 4).value = round(float(g["期初金额"]), 2)
            ws.cell(row, 5).value = round(float(g["入库数量"]), 2)
            ws.cell(row, 6).value = round(float(g["入库金额"]), 2)
            ws.cell(row, 7).value = round(float(g["出库数量"]), 2)
            ws.cell(row, 8).value = round(float(g["出库金额"]), 2)
        else:
            for col in range(3, 9):
                ws.cell(row, col).value = 0

        for col in range(3, 9):
            ws.cell(row, col).number_format = FMT_NUMBER

        if ws.cell(row, 9).value is None:
            ws.cell(row, 9).value  = f"=C{row}+E{row}-G{row}"
        if ws.cell(row, 10).value is None:
            ws.cell(row, 10).value = f"=D{row}+F{row}-H{row}"

        ws.cell(row, 11).value = f"=I{row}"
        ws.cell(row, 12).value = f"=J{row}"


def pandian_build_detail_sheet(wb, sheet_name, category_cn, display_name, df_cat, month):
    """
    新建一张明细表（先删后建），结构与模板「米面粮明细表」完全一致。
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        print(f"  🗑  删除旧 sheet：{sheet_name}")

    ws = wb.create_sheet(title=sheet_name)
    print(f"  📋 新建 sheet：{sheet_name}")

    col_widths = {
        "A": 6.7,  "B": 19.6,
        "C": 9.6,  "D": 9.6,  "E": 11.9,
        "F": 9.6,  "G": 9.6,  "H": 11.9,
        "I": 9.6,  "J": 9.6,  "K": 11.9,
        "L": 9.6,  "M": 9.6,  "N": 11.9,
        "O": 9.6,  "P": 9.6,  "Q": 11.9,
        "R": 9.6,  "S": 9.6,  "T": 9.6,
    }
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w

    ws.row_dimensions[1].height = 33.0
    ws.row_dimensions[2].height = 21.0
    ws.row_dimensions[3].height = 18.0
    ws.row_dimensions[4].height = 18.0

    # 第1行：大标题
    ws.merge_cells("A1:T1")
    c1 = ws["A1"]
    c1.value     = "华东师大泰达附属学校食堂存货盘点汇总表"
    c1.font      = mk_font(bold=True, size=16)
    c1.alignment = mk_align(h="center")

    # 第2行：盘点期间 / 盘点类别 / 单位
    ws.merge_cells("A2:C2")
    ws["A2"].value     = f"盘点期间：2026年{month}月"
    ws["A2"].font      = mk_font(size=12)
    ws["A2"].alignment = mk_align(h="left")
    ws.merge_cells("E2:K2")
    ws["E2"].value     = f"盘点类别：{display_name}"
    ws["E2"].font      = mk_font(size=12)
    ws["E2"].alignment = mk_align(h="center")
    ws["T2"].value     = "单位：元"
    ws["T2"].font      = mk_font(size=12)
    ws["T2"].alignment = mk_align(h="right")

    # 第3行：大组标题辅助函数
    def set_group_header(ws, row, start_col, end_col, label):
        col_range = range(start_col, end_col + 1)
        total = end_col - start_col + 1
        ws.merge_cells(
            start_row=row, start_column=start_col,
            end_row=row, end_column=end_col
        )
        for i, col in enumerate(col_range):
            cell = ws.cell(row, col)
            cell.font      = mk_font(size=12)
            cell.alignment = mk_align(h="center")
            is_first = (i == 0)
            is_last  = (i == total - 1)
            left   = _THIN if is_first else _NONE
            right  = _THIN if is_last  else _NONE
            if total == 1:
                left = right = _THIN
            cell.border = Border(left=left, right=right, top=_THIN, bottom=_THIN)
        ws.cell(row, start_col).value = label

    ws.merge_cells("A3:A4")
    ws.cell(3, 1).value     = "序号"
    ws.cell(3, 1).font      = mk_font(size=12)
    ws.cell(3, 1).alignment = mk_align(h="center")
    ws.cell(3, 1).border    = thin_border()

    ws.merge_cells("B3:B4")
    ws.cell(3, 2).value     = "名称"
    ws.cell(3, 2).font      = mk_font(size=12)
    ws.cell(3, 2).alignment = mk_align(h="center")
    ws.cell(3, 2).border    = thin_border()

    set_group_header(ws, 3, 3,  5,  "期初情况")
    set_group_header(ws, 3, 6,  8,  "本月入库情况")
    set_group_header(ws, 3, 9,  11, "本月出库情况")
    set_group_header(ws, 3, 12, 14, "账面期末情况")
    set_group_header(ws, 3, 15, 17, "盘点情况")
    set_group_header(ws, 3, 18, 20, "盘亏/盘盈情况")

    # 第4行：子列标题
    for col in (1, 2):
        c = ws.cell(4, col)
        c.font      = mk_font(size=12)
        c.alignment = mk_align(h="center")
        c.border    = Border(left=_THIN, right=_THIN, top=_NONE, bottom=_THIN)

    sub_labels = ["数量", "单价", "金额"] * 6
    for i, label in enumerate(sub_labels):
        pandian_sc(ws, 4, 3 + i, label, border=True)

    # 第5行起：数据行
    valid = df_cat[
        (df_cat["期初数量"] != 0) | (df_cat["期初金额"] != 0) |
        (df_cat["入库数量"] != 0) | (df_cat["入库金额"] != 0) |
        (df_cat["出库数量"] != 0) | (df_cat["出库金额"] != 0)
    ].reset_index(drop=True)

    data_start = 5
    n          = len(valid)

    for seq, (_, row_data) in enumerate(valid.iterrows(), 1):
        r = data_start + seq - 1
        ws.row_dimensions[r].height = 18

        pandian_sc(ws, r, 1, seq)
        pandian_sc(ws, r, 2, row_data["货品名称"], h="left")

        c_val = row_data["期初数量"] if row_data["期初数量"] != 0 else ""
        pandian_sc(ws, r, 3, c_val, fmt=FMT_NUMBER if c_val != "" else None)
        pandian_sc(ws, r, 4, f'=IFERROR(E{r}/C{r},"")', fmt=FMT_NUMBER)
        e_val = row_data["期初金额"] if row_data["期初金额"] != 0 else ""
        pandian_sc(ws, r, 5, e_val, fmt=FMT_NUMBER if e_val != "" else None)

        f_val = row_data["入库数量"] if row_data["入库数量"] != 0 else ""
        pandian_sc(ws, r, 6, f_val, fmt=FMT_NUMBER if f_val != "" else None)
        pandian_sc(ws, r, 7, f'=IFERROR(H{r}/F{r},"")', fmt=FMT_NUMBER)
        h_val = row_data["入库金额"] if row_data["入库金额"] != 0 else ""
        pandian_sc(ws, r, 8, h_val, fmt=FMT_NUMBER if h_val != "" else None)

        i_val = row_data["出库数量"] if row_data["出库数量"] != 0 else ""
        pandian_sc(ws, r, 9, i_val, fmt=FMT_NUMBER if i_val != "" else None)
        pandian_sc(ws, r, 10, f'=IFERROR(K{r}/I{r},"")', fmt=FMT_NUMBER)
        k_val = row_data["出库金额"] if row_data["出库金额"] != 0 else ""
        pandian_sc(ws, r, 11, k_val, fmt=FMT_NUMBER if k_val != "" else None)

        pandian_sc(ws, r, 12, f"=N(C{r})+N(F{r})-N(I{r})", fmt=FMT_NUMBER)
        pandian_sc(ws, r, 13, f'=IFERROR(N{r}/L{r},"")', fmt=FMT_NUMBER)
        pandian_sc(ws, r, 14, f"=N(E{r})+N(H{r})-N(K{r})", fmt=FMT_NUMBER)

        pandian_sc(ws, r, 15, f"=L{r}", fmt=FMT_NUMBER)
        pandian_sc(ws, r, 16, f'=IFERROR(Q{r}/O{r},"")', fmt=FMT_NUMBER)
        pandian_sc(ws, r, 17, f"=N{r}", fmt=FMT_NUMBER)

        for col in (18, 19, 20):
            pandian_sc(ws, r, col, None, fmt=FMT_NUMBER)

    # 合计行
    total_row = data_start + n
    data_end  = data_start + n - 1
    ws.row_dimensions[total_row].height = 18

    pandian_sc(ws, total_row, 1, "合计", bold=True)
    pandian_sc(ws, total_row, 2, None, bold=True)

    sum_cols = {
        3:  f"=SUM(C{data_start}:C{data_end})",
        5:  f"=SUM(E{data_start}:E{data_end})",
        6:  f"=SUM(F{data_start}:F{data_end})",
        8:  f"=SUM(H{data_start}:H{data_end})",
        9:  f"=SUM(I{data_start}:I{data_end})",
        11: f"=SUM(K{data_start}:K{data_end})",
        12: f"=SUM(L{data_start}:L{data_end})",
        14: f"=SUM(N{data_start}:N{data_end})",
        15: f"=SUM(O{data_start}:O{data_end})",
        17: f"=SUM(Q{data_start}:Q{data_end})",
    }
    for col in range(3, 21):
        val = sum_cols.get(col, None)
        pandian_sc(ws, total_row, col, val, bold=True, fmt=FMT_NUMBER)

    # 签名行
    sign_row1 = total_row + 2
    sign_row3 = total_row + 5

    for r in (sign_row1, sign_row3):
        ws.row_dimensions[r].height = 17.6

    ws.merge_cells(f"A{sign_row1}:C{sign_row1}")
    ws[f"A{sign_row1}"].value     = "经手人："
    ws[f"A{sign_row1}"].font      = mk_font(size=12)
    ws[f"A{sign_row1}"].alignment = mk_align(h="center")

    ws[f"F{sign_row1}"].value     = "验收人："
    ws[f"F{sign_row1}"].font      = mk_font(size=12)
    ws[f"F{sign_row1}"].alignment = mk_align(h="left")

    ws.merge_cells(f"A{sign_row3}:C{sign_row3}")
    ws[f"A{sign_row3}"].value     = "   食堂负责人："
    ws[f"A{sign_row3}"].font      = mk_font(size=12)
    ws[f"A{sign_row3}"].alignment = mk_align(h="center")


def process_pandian(month):
    """处理指定月份的存货盘点表。"""
    tfile = pandian_target_file(month)

    if not os.path.exists(SOURCE_FILE):
        print(f"❌ 数据源文件不存在：{SOURCE_FILE}")
        return
    if not os.path.exists(tfile):
        print(f"❌ 目标文件不存在：{tfile}（请先准备好该月份的模板文件）")
        return

    print(f"📂 [盘点表] 读取数据源：{SOURCE_FILE}  →  Sheet: 汇总")
    df = load_summary_data()
    print(f"✅ 读取到 {len(df)} 条记录，涉及类别：{sorted(df['货品系列'].unique())}")

    print(f"📂 打开目标文件：{tfile}")
    wb = load_workbook(tfile)

    if "汇总表" not in wb.sheetnames:
        print("❌ 目标文件中不存在[汇总表]，请检查模板文件。")
        return

    print("  📊 更新汇总表...")
    pandian_update_summary_sheet(wb["汇总表"], df, month)

    for sheet_name, cat_cn, display_name in DETAIL_SHEETS:
        df_cat = df[df["货品系列"] == cat_cn].copy()
        print(f"  📋 处理 {sheet_name}（{cat_cn}，共 {len(df_cat)} 条）...")
        pandian_build_detail_sheet(wb, sheet_name, cat_cn, display_name, df_cat, month)

    wb.save(tfile)
    print(f"\n✅ [盘点表] 完成！已保存：{tfile}")


# ══════════════════════════════════════════════════════════════
# 程序入口
# ══════════════════════════════════════════════════════════════

def interactive_mode():
    """交互式菜单，供用户选择功能和月份。"""
    print("\n" + "═" * 50)
    print("  华东食堂出入库及盘点表一体化工具")
    print("═" * 50)
    print("  1. 入库单")
    print("  2. 出库单")
    print("  3. 存货盘点表")
    print("  4. 全部处理（入库 + 出库 + 盘点）")
    print("  0. 退出")
    print("═" * 50)

    while True:
        try:
            choice = input("请选择功能 (0~4) > ").strip()
            if choice == "0":
                print("已退出。")
                return
            if choice in ("1", "2", "3", "4"):
                break
            print("请输入 0~4 之间的数字。")
        except (ValueError, KeyboardInterrupt):
            print("\n已中断。")
            return

    print("请输入要处理的月份（1~12）：")
    while True:
        try:
            m_input = input("月份 > ").strip()
            m = int(m_input)
            if 1 <= m <= 12:
                break
            print("请输入 1~12 之间的整数。")
        except ValueError:
            print("无效输入，请输入数字。")
        except KeyboardInterrupt:
            print("\n已中断。")
            return

    if choice == "1":
        process_ruku(m)
    elif choice == "2":
        process_chuku(m)
    elif choice == "3":
        process_pandian(m)
    elif choice == "4":
        process_ruku(m)
        process_chuku(m)
        process_pandian(m)


def main():
    """
    命令行入口。

    用法示例：
        python canteen_tool.py                      # 交互式菜单
        python canteen_tool.py ruku    --month 1
        python canteen_tool.py chuku   --month 1
        python canteen_tool.py pandian --month 1
        python canteen_tool.py all     --month 1    # 三个模块一起处理
    """
    parser = argparse.ArgumentParser(
        description="华东食堂出入库及盘点表一体化自动填写工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
子命令：
  ruku      处理入库单
  chuku     处理出库单
  pandian   处理存货盘点表
  all       同时处理三个模块

示例：
  python canteen_tool.py ruku --month 1
  python canteen_tool.py all  --month 3
        """
    )
    parser.add_argument(
        "mode",
        nargs="?",
        choices=["ruku", "chuku", "pandian", "all"],
        help="功能模块（不传则进入交互模式）"
    )
    parser.add_argument("--month", type=int, help="要处理的月份（1~12）")
    args = parser.parse_args()

    if args.mode is None:
        # 无参数 → 交互模式
        interactive_mode()
        return

    if args.month is None:
        parser.error("使用子命令时必须同时指定 --month 参数，例如：--month 1")

    if args.mode == "ruku":
        process_ruku(args.month)
    elif args.mode == "chuku":
        process_chuku(args.month)
    elif args.mode == "pandian":
        process_pandian(args.month)
    elif args.mode == "all":
        process_ruku(args.month)
        process_chuku(args.month)
        process_pandian(args.month)


if __name__ == "__main__":
    main()